import os
import json
import sqlite3
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
import pandas as pd
from telegram import Update, KeyboardButton, ReplyKeyboardMarkup, WebAppInfo
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, filters, ContextTypes

TOKEN = "8847383930:AAH3O_2GC9x-iERPfr7FqdiX8zzwMhSnqVA"
WEBAPP_URL = "https://frtnbura.github.io/firtina-lojistik/"  # HTML dosyanızın HTTPS linki
GITHUB_TOKEN = "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"
GITHUB_REPO = "frtnbura/firtina-lojistik"

# Mevcut Masaüstü Dosyaları
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_DOSYA = os.path.join(BASE_DIR, "nakliye_kamyon_hesap_takip.xlsx")
EXCEL_DIKILI = os.path.join(BASE_DIR, "dikili_sefer_hesap_takip.xlsx")
EXCEL_GIDERLER = os.path.join(BASE_DIR, "sabit_giderler_maaslar.xlsx")
EXCEL_ARIZA = os.path.join(BASE_DIR, "ariza_sanayi_bakim.xlsx")
DB_YAKIT = os.path.join(BASE_DIR, "yakit_takip.db")

KATEGORILER = {
    "Yüklemeciler": os.path.join(BASE_DIR, "yevmiye_yuklemeciler.xlsx"),
    "Şoförler": os.path.join(BASE_DIR, "yevmiye_soforler.xlsx"),
    "Kesimciler": os.path.join(BASE_DIR, "yevmiye_kesimciler.xlsx"),
    "Diğer": os.path.join(BASE_DIR, "yevmiye_diger.xlsx")
}

AYLAR_TR = ["TÜMÜ", "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran", "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"]

# --- TELEGRAM START ---
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    kb = [[KeyboardButton(text="⚡ FIRTINA LOJİSTİK YÖNETİM", web_app=WebAppInfo(url=WEBAPP_URL))]]
    reply_markup = ReplyKeyboardMarkup(kb, resize_keyboard=True)
    await update.message.reply_text(
        "🚛 *FIRTINA LOJİSTİK - MOBİL ENTEGRASYON*\n\n"
        "Aşağıdaki butona basarak Sefer, Dikili, Veresiye, Eleman/Yevmiye ve Sanayi masraflarını doğrudan bilgisayardaki programa işleyebilirsiniz.",
        reply_markup=reply_markup,
        parse_mode="Markdown"
    )

# --- TELEGRAM WEBAPP VERİLERİNİ EXCEL VE SQLITE'A İŞLEME ---
async def web_app_veri_yakala(update: Update, context: ContextTypes.DEFAULT_TYPE):
    veri = json.loads(update.effective_message.web_app_data.data)
    modul = veri.get("modul")
    mesaj = "İşlem kaydedildi."

    # 1. NAKLİYE SEFER MODÜLÜ
    if modul == "nakliye_sefer":
        tutar = veri["tonaj"] * veri["fiyat"]
        kdv = tutar * 0.20
        tevkifat = kdv * 0.20
        toplam = tutar + (kdv - tevkifat)
        kalan_kar = toplam - (veri["yakit"] + veri["yukleme"] + veri["harcirah"])

        wb = load_workbook(EXCEL_DOSYA)
        ws = wb["Sefer Kayıtları"]
        ws.append([
            veri["tarih"], veri["plaka"], veri["sofor"], veri["tonaj"], veri["fiyat"],
            tutar, kdv, tevkifat, toplam, veri["rampa"], veri["yakit"], veri["yukleme"],
            veri["harcirah"], kalan_kar
        ])
        wb.save(EXCEL_DOSYA)

        mesaj = (
            f"✅ *NAKLİYE SEFERİ PC'YE EKLENDİ!*\n"
            f"📅 Tarih: `{veri['tarih']}` | 🚛 Plaka: `{veri['plaka']}`\n"
            f"⚖️ Tonaj: *{veri['tonaj']:,.2f} Ton* | 💰 Net Kâr: *{kalan_kar:,.2f} TL*"
        )

    # 2. DİKİLİ SEFER MODÜLÜ
    elif modul == "dikili_sefer":
        wb = load_workbook(EXCEL_DIKILI)
        ws = wb["Dikili_Seferler"]
        ws.append([
            veri["tarih"], veri["plaka"], veri["sofor"], veri["kesimci"],
            veri["tonaj"], veri["ster"], veri["tur"]
        ])
        wb.save(EXCEL_DIKILI)

        mesaj = (
            f"🌲 *DİKİLİ SEFERİ EKLENDİ!*\n"
            f"📅 `{veri['tarih']}` | 🚛 `{veri['plaka']}`\n"
            f"⚖️ {veri['tonaj']:,.2f} Ton / {veri['ster']} Ster ({veri['tur']})"
        )

    # 3. DİKİLİ VERESİYE MODÜLÜ
    elif modul == "dikili_veresiye":
        toplam_tutar = veri["tonaj"] * veri["fiyat"]
        kalan_borc = toplam_tutar - veri["odeme"]
        wb = load_workbook(EXCEL_DIKILI)
        ws = wb["Dikili_Veresiye"]
        vid = ws.max_row
        ws.append([
            vid, veri["tarih"], veri["musteri"], veri["tonaj"], veri["fiyat"],
            toplam_tutar, veri["odeme"], kalan_borc, veri["aciklama"]
        ])
        wb.save(EXCEL_DIKILI)

        mesaj = (
            f"💳 *DİKİLİ VERESİYE İŞLENDİ!*\n"
            f"👤 Müşteri: *{veri['musteri']}*\n"
            f"💵 Toplam: {toplam_tutar:,.2f} TL | Tahsilat: {veri['odeme']:,.2f} TL\n"
            f"📌 Kalan Borç: *{kalan_borc:,.2f} TL*"
        )

    # 4. ELEMAN & YEVMİYE MODÜLÜ
    elif modul == "eleman_yevmiye":
        dosya = KATEGORILER.get(veri["kategori"], KATEGORILER["Diğer"])
        df_eski = pd.read_excel(dosya, sheet_name="Yevmiye Geçmişi")
        kimlik = int(df_eski["Kimlik"].max() + 1) if len(df_eski) > 0 and not pd.isna(df_eski["Kimlik"].max()) else 1

        yeni_satir = pd.DataFrame([{
            "Kimlik": kimlik, "Tarih": veri["tarih"], "Eleman Adı": veri["eleman"],
            "İşlem Türü": veri["tur"], "Miktar (TL)": veri["miktar"], "Açıklama": veri["aciklama"]
        }])
        df_yeni = pd.concat([df_eski, yeni_satir], ignore_index=True)

        # Özet güncelleme
        elemanlar = df_yeni["Eleman Adı"].unique()
        ozet = []
        for e in elemanlar:
            e_df = df_yeni[df_yeni["Eleman Adı"] == e]
            t_y = e_df[e_df["İşlem Türü"] == "Yevmiye Ekle"]["Miktar (TL)"].sum()
            t_a = e_df[e_df["İşlem Türü"] == "Avans/Ödeme Düş"]["Miktar (TL)"].sum()
            ozet.append({"Eleman Adı": e, "Toplam Hak Edilen": t_y, "Toplam Ödenen/Avans": t_a, "Kalan Alacak (TL)": t_y - t_a})

        with pd.ExcelWriter(dosya, engine='openpyxl') as writer:
            df_yeni.to_excel(writer, sheet_name="Yevmiye Geçmişi", index=False)
            pd.DataFrame(ozet).to_excel(writer, sheet_name="Eleman Özet Listesi", index=False)

        mesaj = (
            f"👷 *YEVMİYE KAYDEDİLDİ!*\n"
            f"🏷️ Kategori: {veri['kategori']} | 👤 Eleman: *{veri['eleman']}*\n"
            f"📝 İşlem: {veri['tur']} | 💵 Tutar: *{veri['miktar']:,.2f} TL*"
        )

    # 5. YAKIT MODÜLÜ
    elif modul == "yakit":
        conn = sqlite3.connect(DB_YAKIT)
        cur = conn.cursor()
        cur.execute("INSERT INTO yakit_kayitlari (plaka, tarih, litre, tutar, sofor) VALUES (?, ?, ?, ?, ?)",
                    (veri["plaka"], veri["tarih"], veri["litre"], veri["tutar"], veri["sofor"]))
        conn.commit()
        conn.close()

        mesaj = f"⛽ *YAKIT İŞLENDİ:* `{veri['plaka']}` - {veri['litre']} Litre (*{veri['tutar']:,.2f} TL*)"

    # 6. SABİT GİDER & SİGORTA MODÜLÜ
    elif modul == "sabit_gider":
        dt = datetime.strptime(veri["tarih"], "%d.%m.%Y")
        yil = str(dt.year)
        wb = load_workbook(EXCEL_GIDERLER)
        ws = wb["Sabit_Giderler"]
        gid = ws.max_row
        ws.append([gid, veri["tarih"], yil, veri["tur"], veri["ilgili"], veri["tutar"], veri["aciklama"]])
        wb.save(EXCEL_GIDERLER)

        mesaj = f"🛡️ *GİDER İŞLENDİ:* {veri['tur']} ({veri['ilgili']}) - *{veri['tutar']:,.2f} TL*"

    # 7. ARIZA & SANAYİ MODÜLÜ
    elif modul == "ariza_sanayi":
        dt = datetime.strptime(veri["tarih"], "%d.%m.%Y")
        yil = str(dt.year)
        ay = AYLAR_TR[dt.month]
        wb = load_workbook(EXCEL_ARIZA)
        ws = wb["Ariza_Bakim"]
        aid = ws.max_row
        ws.append([aid, veri["tarih"], yil, ay, veri["plaka"], veri["tur"], veri["tutar"], veri["servis"], veri["islem"]])
        wb.save(EXCEL_ARIZA)

        mesaj = f"🛠️ *SANAYİ MASRAFI İŞLENDİ:* `{veri['plaka']}` - {veri['tur']} (*{veri['tutar']:,.2f} TL*)"

    await update.message.reply_text(mesaj, parse_mode="Markdown")

if __name__ == "__main__":
    app = ApplicationBuilder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.StatusUpdate.WEB_APP_DATA, web_app_veri_yakala))
    print("Fırtına Lojistik Telegram Köprüsü Başlatıldı...")
    app.run_polling()
